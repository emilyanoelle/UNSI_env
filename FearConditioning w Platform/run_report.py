# -*- coding: utf-8 -*-
"""
run_report.py
-------------
Generates / updates a single Excel run report workbook with one sheet per
analysis type.  Uses a read-merge-write strategy: if run_report.xlsx already
exists, only the sheets for analyses that ran this session are replaced.
Sheets for analyses that did NOT run this session are left exactly as they
were, including their timestamps.

Output: <ANALYSIS_OUTPUT_DIR>/run_report.xlsx
"""

import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Colours / style constants ───────────────────────────────────────────────

_C = {
    "header_dark":  "FF1F3864",
    "header_mid":   "FF2E5FA3",
    "header_light": "FFD6E4F7",
    "ok":           "FFE2EFDA",
    "warn":         "FFFFF2CC",
    "error":        "FFFCE4D6",
    "white":        "FFFFFFFF",
    "alt":          "FFF5F9FF",
    "ts_fresh":     "FF375623",   # dark green  — updated this run
    "ts_stale":     "FF595959",   # dark grey   — not updated this run
    "never":        "FF7F3F00",   # dark orange — never run
}

_THIN   = Side(style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Canonical analysis keys → human-readable sheet names
ANALYSIS_KEYS = {
    "freezing":         "% Time Freezing",
    "freezing_bouts":   "Freezing Bouts",
    "platform":         "Platform Time",
    "platform_latency": "Platform Latency",
    "us_locked":        "US Locked",
    "eee":              "EEE",
    "event_raster":     "Event Rasters",
    "speed":            "Speed",
}

# Sheet name → tab colour
_TAB_COLORS = {
    "Overview":           "1F3864",
    "% Time Freezing":    "375623",
    "Freezing Bouts":     "375623",
    "Platform Time":      "833C00",
    "Platform Latency":   "833C00",
    "US Locked":          "4472C4",
    "EEE":                "7030A0",
    "Event Rasters":      "C65911",
    "Speed":              "1F6B3A",
    "Excluded Subjects":  "C00000",
    "Per-Subject Log":    "595959",
}

# Desired sheet order in the final workbook
_SHEET_ORDER = [
    "Overview",
    "% Time Freezing",
    "Freezing Bouts",
    "Platform Time",
    "Platform Latency",
    "US Locked",
    "EEE",
    "Event Rasters",
    "Speed",
    "Excluded Subjects",
    "Per-Subject Log",
]


# ── Public API - called from analysis modules ───────────────────────────────

def new_report(cfg: dict) -> dict:
    return {
        "timestamp":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "python_version": sys.version,
        "cfg":            cfg,
        "logs": defaultdict(lambda: {
            "subject_logs":      {},
            "excluded_subjects": {},
            "n_csv_total":       0,
            "extra_overview":    {},
        }),
    }


def record_subject(report: dict, analysis: str, subject_key: str,
                   columns_used: dict, warnings: Optional[list] = None,
                   skipped: Optional[list] = None):
    log = report["logs"][analysis]
    log["subject_logs"][subject_key] = {
        "columns_used": columns_used,
        "warnings":     warnings or [],
        "skipped":      skipped  or [],
    }
    log["n_csv_total"] += 1


def record_exclusion(report: dict, analysis: str, subject_key: str, reason: str):
    report["logs"][analysis]["excluded_subjects"][subject_key] = reason


def record_overview(report: dict, analysis: str, key: str, value):
    report["logs"][analysis]["extra_overview"][key] = value


# ── openpyxl helpers ────────────────────────────────────────────────────────

def _hdr(ws, row, col, value, bg=_C["header_dark"], fg="FFFFFFFF",
         bold=True, size=11, wrap=False, colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical="center")
    cell.border    = _BORDER
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + colspan - 1)
    return cell


def _cell(ws, row, col, value, bg=_C["white"], bold=False, wrap=False, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical="center")
    cell.border    = _BORDER
    return cell


def _section(ws, row, col, title, ncols=4):
    _hdr(ws, row, col, title, bg=_C["header_mid"], size=10, colspan=ncols)
    return row + 1


def _blank(ws, row):
    ws.row_dimensions[row].height = 6
    return row + 1


def _kv(ws, row, col, items, key_bg=_C["header_light"]):
    items = list(items)
    for k, v in items:
        _cell(ws, row, col,     str(k), bg=key_bg, bold=True)
        _cell(ws, row, col + 1, str(v) if v is not None else "—", wrap=True)
        row += 1
    return row


def _table(ws, row, col, headers, rows, widths=None):
    for ci, h in enumerate(headers, col):
        _hdr(ws, row, ci, h, bg=_C["header_mid"], size=10)
    row += 1
    for ri, rd in enumerate(rows):
        bg = _C["alt"] if ri % 2 else _C["white"]
        for ci, h in enumerate(headers, col):
            _cell(ws, row, ci, rd.get(h, ""), bg=bg, wrap=True)
        row += 1
    if widths:
        for ci, h in enumerate(headers, col):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 18)
    return row


# ── Timestamp banner ────────────────────────────────────────────────────────

def _write_ts_banner(ws, timestamp: str, fresh: bool, ncols: int = 4):
    """
    Row 1 of an analysis sheet: a coloured banner showing when it was last run.
    fresh=True  → green  (updated this session)
    fresh=False → grey   (carried over from a previous run)
    """
    bg  = _C["ts_fresh"] if fresh else _C["ts_stale"]
    lbl = f"Last updated: {timestamp}" if timestamp else "Never run"
    if not timestamp:
        bg = _C["never"]
    _hdr(ws, 1, 1, lbl, bg=bg, size=10, colspan=ncols)
    ws.row_dimensions[1].height = 18


# ── Column overview builder ─────────────────────────────────────────────────

def _col_overview(log: dict) -> list:
    role_counts = defaultdict(lambda: defaultdict(int))
    n_total = max(log["n_csv_total"], len(log["subject_logs"]))
    for slog in log["subject_logs"].values():
        for role, col in slog.get("columns_used", {}).items():
            role_counts[role][str(col)] += 1
    rows = []
    for role, col_counts in role_counts.items():
        best  = max(col_counts, key=col_counts.get)
        n     = col_counts[best]
        other = {k: v for k, v in col_counts.items() if k != best}
        note  = ("also seen: " + "; ".join(
            f"'{k}' in {v} file(s)" for k, v in other.items())) if other else ""
        rows.append({"Column role": role, "Matched column": best,
                     "Found in": f"{n} of {n_total} CSV files", "Notes": note})
    return rows


# ── Existing-file timestamp reader ──────────────────────────────────────────

def _read_existing_timestamps(path: Path) -> dict:
    """
    Read the 'last updated' timestamp from cell A1 of each sheet in an
    existing workbook.  Returns {sheet_name: timestamp_string or None}.
    """
    ts = {}
    if not path.exists():
        return ts
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws  = wb[name]
            val = ws.cell(row=1, column=1).value or ""
            # Banner text is "Last updated: YYYY-MM-DD HH:MM:SS"
            if isinstance(val, str) and val.startswith("Last updated:"):
                ts[name] = val.replace("Last updated:", "").strip()
            else:
                ts[name] = None
        wb.close()
    except Exception:
        pass
    return ts


# ── Sheet builders ──────────────────────────────────────────────────────────

def _overview_sheet(wb, report: dict, existing_ts: dict, ran_this_session: set):
    """
    Overview sheet with a 3-column analysis status table:
        Analysis | Status | Last run timestamp
    """
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26

    cfg = report["cfg"]
    now = report["timestamp"]
    r   = 1

    _hdr(ws, r, 1, "FEAR CONDITIONING PIPELINE — RUN REPORT",
         bg=_C["header_dark"], size=13, colspan=3)
    r += 1
    _cell(ws, r, 1, f"Report opened: {now}", bg=_C["header_light"], bold=True)
    _cell(ws, r, 2, f"Python {report['python_version'].split()[0]}",
          bg=_C["header_light"])
    r = _blank(ws, r + 1)

    # ── Analysis status table ─────────────────────────────────────────────────
    r = _section(ws, r, 1, "Analysis run status", ncols=3)
    _hdr(ws, r, 1, "Analysis",       bg=_C["header_mid"], size=10)
    _hdr(ws, r, 2, "Status",         bg=_C["header_mid"], size=10)
    _hdr(ws, r, 3, "Last run",       bg=_C["header_mid"], size=10)
    r += 1

    analysis_display = [
        ("freezing",         "% Time Freezing",  "% Time Freezing"),
        ("freezing_bouts",   "Freezing Bouts",   "Freezing Bouts"),
        ("platform",         "Platform Time",    "Platform Time"),
        ("platform_latency", "Platform Latency", "Platform Latency"),
        ("us_locked",        "US Locked",        "US Locked"),
        ("eee",              "EEE",              "EEE"),
        ("event_raster",     "Event Rasters",    "Event Rasters"),
        ("speed",            "Speed",            "Speed"),
    ]

    for ri, (key, label, sheet_name) in enumerate(analysis_display):
        bg       = _C["alt"] if ri % 2 else _C["white"]
        ran      = sheet_name in ran_this_session
        prev_ts  = existing_ts.get(sheet_name)

        if ran:
            status   = "✓ Run this session"
            ts_shown = now
            status_bg = _C["ok"]
        elif prev_ts:
            status   = "– Not run (previous data kept)"
            ts_shown = prev_ts
            status_bg = _C["warn"]
        else:
            status   = "✗ Never run"
            ts_shown = "—"
            status_bg = _C["error"]

        _cell(ws, r, 1, label,     bg=bg)
        _cell(ws, r, 2, status,    bg=status_bg)
        _cell(ws, r, 3, ts_shown,  bg=bg)
        r += 1

    r = _blank(ws, r)

    # ── Directories ───────────────────────────────────────────────────────────
    r = _section(ws, r, 1, "Directories", ncols=3)
    for i, bd in enumerate(cfg.get("behaviordata_dirs", []), 1):
        _cell(ws, r, 1, f"BehaviorData dir {i}", bg=_C["header_light"], bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        _cell(ws, r, 2, str(bd), wrap=True)
        r += 1
    _cell(ws, r, 1, "Analysis output dir", bg=_C["header_light"], bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    _cell(ws, r, 2, str(cfg.get("analysis_out", "—")), wrap=True)
    r = _blank(ws, r + 1)

    # ── Treatment groups ──────────────────────────────────────────────────────
    r = _section(ws, r, 1, "Treatment groups", ncols=3)
    for grp in cfg.get("canonical_groups", []):
        color = cfg.get("treatment_colors", {}).get(grp, "")
        _cell(ws, r, 1, grp,            bg=_C["header_light"], bold=True)
        _cell(ws, r, 2, f"color: {color}")
        r += 1
    r = _blank(ws, r)

    # ── Key settings ──────────────────────────────────────────────────────────
    r = _section(ws, r, 1, "Key settings", ncols=3)
    settings = [
        ("CS detection mode",        cfg.get("cs_detection_mode")),
        ("Column match mode",        cfg.get("column_match_mode")),
        ("CS+ tone-status pattern",  cfg.get("tone_status_col_csplus")),
        ("CS- tone-status pattern",  cfg.get("tone_status_col_csminus")),
        ("CS trial cap",             cfg.get("cs_trial_cap")),
        ("ITI trial cap",            cfg.get("iti_trial_cap")),
        ("EEE trial cap",            cfg.get("eee_trial_cap")),
        ("CSV suffix filter",        repr(cfg.get("csv_suffix", ""))),
        ("Prism export",             cfg.get("prism_export")),
        ("Use shocker column (US)",  cfg.get("use_shocker_column")),
        ("US duration (s)",          cfg.get("us_duration_s")),
        ("US chance baseline (%)",   cfg.get("us_chance_baseline_pct")),
        ("Include treatments (US)",  cfg.get("include_treatments") or "all"),
        ("Exclude behavior IDs",     cfg.get("exclude_behavior_ids") or "none"),
        ("Heatmap sort",             cfg.get("heatmap_sort")),
        ("Freezing by sex",          cfg.get("freezing_by_sex")),
        ("Freezing by litter",       cfg.get("freezing_by_litter")),
        ("Platform by sex",          cfg.get("platform_by_sex")),
        ("EEE by sex",               cfg.get("eee_by_sex")),
        ("Event raster subfolder",   cfg.get("event_raster_subfolder")),
    ]
    for k, v in settings:
        _cell(ws, r, 1, str(k), bg=_C["header_light"], bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        _cell(ws, r, 2, str(v) if v is not None else "—", wrap=True)
        r += 1


def _analysis_sheet(wb, sheet_name: str, analysis_key: str, report: dict,
                    timestamp: str, fresh: bool,
                    extra_settings: Optional[list] = None):
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [28, 38, 22, 40]):
        ws.column_dimensions[col].width = w

    log = report["logs"][analysis_key]

    # Row 1: timestamp banner
    _write_ts_banner(ws, timestamp, fresh, ncols=4)

    r = 2
    _hdr(ws, r, 1, sheet_name.upper(), bg=_C["header_dark"], size=12, colspan=4)
    r = _blank(ws, r + 1)

    # 1. Column overview
    r = _section(ws, r, 1,
        f"Column overview  ({log['n_csv_total']} CSV files processed)", ncols=4)
    col_rows = _col_overview(log)
    if col_rows:
        r = _table(ws, r, 1,
            headers=["Column role", "Matched column", "Found in", "Notes"],
            rows=col_rows,
            widths={"Column role": 26, "Matched column": 36,
                    "Found in": 22, "Notes": 38})
    else:
        _cell(ws, r, 1, "(no column data recorded — analysis may not have run)",
              bg=_C["warn"])
        r += 1
    r = _blank(ws, r)

    # 2. Extra settings
    combined = list(extra_settings or []) + list(log["extra_overview"].items())
    if combined:
        r = _section(ws, r, 1, "Analysis-specific settings", ncols=4)
        r = _kv(ws, r, 1, combined)
        r = _blank(ws, r)

    # 3. Exclusions
    excl = log["excluded_subjects"]
    r = _section(ws, r, 1, f"Excluded subjects  ({len(excl)} total)", ncols=4)
    if excl:
        excl_rows = [{"File": k, "Reason": v} for k, v in excl.items()]
        r = _table(ws, r, 1,
            headers=["File", "Reason"],
            rows=excl_rows,
            widths={"File": 48, "Reason": 52})
        for ri in range(r - len(excl_rows), r):
            ws.cell(ri, 2).fill = PatternFill("solid", fgColor=_C["error"])
    else:
        _cell(ws, r, 1, "(none)", bg=_C["ok"])
        r += 1
    r = _blank(ws, r)

    # 4. Per-subject log
    slogs = log["subject_logs"]
    r = _section(ws, r, 1,
        f"Per-subject validation log  ({len(slogs)} subjects)", ncols=4)
    flat = []
    for subj, slog in slogs.items():
        issues = "; ".join(
            [f"WARN: {w}" for w in slog.get("warnings", [])] +
            [f"SKIP: {s}" for s in slog.get("skipped",   [])]
        )
        flat.append({
            "File":         subj,
            "Columns used": "  |  ".join(
                f"{k}: {v}" for k, v in slog.get("columns_used", {}).items()),
            "Issues":       issues or "—",
        })
    if flat:
        r = _table(ws, r, 1,
            headers=["File", "Columns used", "Issues"],
            rows=flat,
            widths={"File": 40, "Columns used": 60, "Issues": 48})
        data_start = r - len(flat)
        for ri, fd in enumerate(flat):
            if fd["Issues"] != "—":
                ws.cell(data_start + ri, 3).fill = \
                    PatternFill("solid", fgColor=_C["warn"])
    else:
        _cell(ws, r, 1, "(no subjects processed)", bg=_C["warn"])


def _exclusions_sheet(wb, report: dict, timestamp: str):
    ws = wb.create_sheet("Excluded Subjects")
    ws.sheet_view.showGridLines = False

    _write_ts_banner(ws, timestamp, fresh=True, ncols=3)
    r = 2
    _hdr(ws, r, 1, "MASTER EXCLUSION TABLE — ALL ANALYSES",
         bg=_C["header_dark"], size=12, colspan=3)
    r = _blank(ws, r + 1)

    rows = []
    for key, label in ANALYSIS_KEYS.items():
        for subj, reason in report["logs"][key]["excluded_subjects"].items():
            rows.append({"Analysis": label, "File": subj, "Reason": reason})
    if rows:
        r = _table(ws, r, 1,
            headers=["Analysis", "File", "Reason"],
            rows=rows,
            widths={"Analysis": 22, "File": 50, "Reason": 55})
        for ri in range(3, r):
            ws.cell(ri, 3).fill = PatternFill("solid", fgColor=_C["error"])
    else:
        _cell(ws, r, 1, "(no subjects excluded across any analysis)", bg=_C["ok"])


def _persubject_sheet(wb, report: dict, timestamp: str):
    ws = wb.create_sheet("Per-Subject Log")
    ws.sheet_view.showGridLines = False

    _write_ts_banner(ws, timestamp, fresh=True, ncols=4)
    r = 2
    _hdr(ws, r, 1, "MASTER PER-SUBJECT COLUMN-MATCH LOG — ALL ANALYSES",
         bg=_C["header_dark"], size=12, colspan=4)
    r = _blank(ws, r + 1)

    rows = []
    for key, label in ANALYSIS_KEYS.items():
        for subj, slog in report["logs"][key]["subject_logs"].items():
            issues = "; ".join(
                [f"WARN: {w}" for w in slog.get("warnings", [])] +
                [f"SKIP: {s}" for s in slog.get("skipped",   [])]
            )
            rows.append({
                "Analysis":     label,
                "File":         subj,
                "Columns used": "  |  ".join(
                    f"{k}: {v}" for k, v in slog.get("columns_used", {}).items()),
                "Issues":       issues or "—",
            })
    if rows:
        r = _table(ws, r, 1,
            headers=["Analysis", "File", "Columns used", "Issues"],
            rows=rows,
            widths={"Analysis": 20, "File": 42,
                    "Columns used": 65, "Issues": 50})
        data_start = 3
        for ri, rd in enumerate(rows):
            if rd["Issues"] != "—":
                ws.cell(data_start + ri, 4).fill = \
                    PatternFill("solid", fgColor=_C["warn"])
    else:
        _cell(ws, r, 1, "(no subject logs recorded)", bg=_C["warn"])


# ── Main entry point ────────────────────────────────────────────────────────

def write_excel_report(report: dict, out_dir: Path):
    """
    Build or update run_report.xlsx using read-merge-write:
      - Sheets for analyses run this session are fully replaced.
      - Sheets for analyses NOT run this session are carried over unchanged.
      - Overview and master sheets are always regenerated.
    """
    cfg      = report["cfg"]
    now      = report["timestamp"]
    out_path = out_dir / "run_report.xlsx"

    # Read timestamps from any existing workbook before we touch anything
    existing_ts = _read_existing_timestamps(out_path)

    # Work out which analysis sheets will be regenerated this session
    ran_this_session = set()   # sheet names rebuilt this run

    # Decide which analysis sheets to write this run
    sheets_to_write = {}   # sheet_name → (analysis_key, extra_settings)

    if cfg.get("run_freezing"):
        sheets_to_write["% Time Freezing"] = (
            "freezing",
            [("Subfolder",    cfg.get("freezing_subfolder")),
             ("CS trial cap", cfg.get("cs_trial_cap")),
             ("ITI trial cap",cfg.get("iti_trial_cap")),
             ("By sex",       cfg.get("freezing_by_sex")),
             ("By litter",    cfg.get("freezing_by_litter")),
             ("Prism export", cfg.get("prism_export"))],
        )
    if cfg.get("run_freezing") and cfg.get("freezing_bouts"):
        sheets_to_write["Freezing Bouts"] = (
            "freezing_bouts",
            [("Subfolder",    cfg.get("freezing_subfolder")),
             ("CS trial cap", cfg.get("cs_trial_cap")),
             ("ITI trial cap",cfg.get("iti_trial_cap")),
             ("Prism export", cfg.get("prism_export"))],
        )
    if cfg.get("run_platform"):
        sheets_to_write["Platform Time"] = (
            "platform",
            [("Subfolder",    cfg.get("platform_subfolder")),
             ("CS trial cap", cfg.get("cs_trial_cap")),
             ("ITI trial cap",cfg.get("iti_trial_cap")),
             ("By sex",       cfg.get("platform_by_sex")),
             ("Prism export", cfg.get("prism_export"))],
        )
    if cfg.get("run_platform") and cfg.get("platform_latency"):
        sheets_to_write["Platform Latency"] = (
            "platform_latency",
            [("Latency col source",
              "AnyMaze export column if found, else computed from in_platform"),
             ("Subfolder", cfg.get("latency_subfolder"))],
        )
    if cfg.get("run_us_locked"):
        mode = ("Mode A — Shocker active column" if cfg.get("use_shocker_column")
                else f"Mode B — last {cfg.get('us_duration_s', 2.0):.1f}s of CS+ trial")
        sheets_to_write["US Locked"] = (
            "us_locked",
            [("US window mode",       mode),
             ("US duration (s)",      cfg.get("us_duration_s")),
             ("Chance baseline (%)",  cfg.get("us_chance_baseline_pct")),
             ("Include treatments",   cfg.get("include_treatments") or "all"),
             ("Exclude behavior IDs", cfg.get("exclude_behavior_ids") or "none"),
             ("Heatmap sort",         cfg.get("heatmap_sort")),
             ("CS trial cap",         cfg.get("cs_trial_cap")),
             ("Prism export",         cfg.get("prism_export"))],
        )
    if cfg.get("run_eee"):
        sheets_to_write["EEE"] = (
            "eee",
            [("Subfolder",    cfg.get("eee_subfolder")),
             ("EEE trial cap",cfg.get("eee_trial_cap")),
             ("By sex",       cfg.get("eee_by_sex")),
             ("Prism export", cfg.get("prism_export")),
             ("Outcome logic",
               "evade=100% on platform | endure=0% | escape=mounted during US")],
        )

    if cfg.get("run_event_raster"):
        sheets_to_write["Event Rasters"] = (
            "event_raster",
            [("Subfolder", cfg.get("event_raster_subfolder")),
             ("Output",    "pass-1 per-session SVG only"),
             ("Tracks",    "Freezing | Platform | CS+ | CS- | US"),
             ("US fallback",
              "preferred configured source, then alternate raw source, then last US_DURATION_S of CS+")],
        )

    if cfg.get("run_speed"):
        sheets_to_write["Speed"] = (
            "speed",
            [("Subfolder",       cfg.get("speed_subfolder")),
             ("Pre-onset (s)",   cfg.get("speed_pre_bins",  300) / 10),
             ("Post-onset (s)",  cfg.get("speed_post_bins", 600) / 10),
             ("Bin width",       "100 ms"),
             ("Prism export",    "n/a — outputs are Excel workbooks directly")],
        )

    ran_this_session = set(sheets_to_write.keys())
    # Overview and master sheets are always regenerated
    ran_this_session.update(["Overview", "Excluded Subjects", "Per-Subject Log"])

    # ── Load existing workbook or start fresh ─────────────────────────────────
    if out_path.exists():
        try:
            wb_old = load_workbook(out_path)
            old_sheets = {name: wb_old[name] for name in wb_old.sheetnames}
        except Exception as e:
            print(f"[warn] Could not read existing run_report.xlsx ({e}); starting fresh.")
            wb_old    = None
            old_sheets = {}
    else:
        wb_old     = None
        old_sheets = {}

    # ── Build the new workbook ────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in _SHEET_ORDER:

        if sheet_name == "Overview":
            _overview_sheet(wb, report, existing_ts, ran_this_session)

        elif sheet_name in ("Excluded Subjects", "Per-Subject Log"):
            if sheet_name == "Excluded Subjects":
                _exclusions_sheet(wb, report, now)
            else:
                _persubject_sheet(wb, report, now)

        elif sheet_name in sheets_to_write:
            # Regenerate with fresh data
            analysis_key, extra = sheets_to_write[sheet_name]
            _analysis_sheet(wb, sheet_name, analysis_key, report,
                            timestamp=now, fresh=True, extra_settings=extra)

        elif sheet_name in old_sheets:
            # Carry over the unchanged sheet from the previous workbook
            _copy_sheet(wb_old, wb, sheet_name)

        # else: sheet was never created and didn't run — skip silently

    # ── Tab colours ───────────────────────────────────────────────────────────
    for ws in wb.worksheets:
        c = _TAB_COLORS.get(ws.title)
        if c:
            ws.sheet_properties.tabColor = c

    if wb_old:
        wb_old.close()

    wb.save(out_path)
    print(f"[ok] Run report saved: {out_path}")
    return out_path


# ── Sheet copy helper ───────────────────────────────────────────────────────

def _copy_sheet(wb_src, wb_dst, sheet_name: str):
    """
    Copy a worksheet from wb_src into wb_dst preserving cell values, styles,
    dimensions, and merge info.  Does not copy charts or images.
    """
    from copy import copy as cp
    ws_src = wb_src[sheet_name]
    ws_dst = wb_dst.create_sheet(sheet_name)

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column,
                                   value=cell.value)
            if cell.has_style:
                new_cell.font      = cp(cell.font)
                new_cell.fill      = cp(cell.fill)
                new_cell.border    = cp(cell.border)
                new_cell.alignment = cp(cell.alignment)
                new_cell.number_format = cell.number_format

    for merge in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge))

    for col_letter, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_idx].height = dim.height

    ws_dst.sheet_view.showGridLines = ws_src.sheet_view.showGridLines
